"""
Ekspor laporan emosi ke Excel (.xlsx) — 3 sheet lengkap
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

EMOSI_META = {
    'angry':    {'label': 'Marah',    'emoji': '😠', 'hex': 'FFEF4444'},
    'disgust':  {'label': 'Jijik',    'emoji': '🤢', 'hex': 'FF8B5CF6'},
    'fear':     {'label': 'Takut',    'emoji': '😨', 'hex': 'FF6366F1'},
    'happy':    {'label': 'Senang',   'emoji': '😊', 'hex': 'FF22C55E'},
    'sad':      {'label': 'Sedih',    'emoji': '😢', 'hex': 'FF3B82F6'},
    'surprise': {'label': 'Terkejut', 'emoji': '😲', 'hex': 'FFF59E0B'},
    'neutral':  {'label': 'Netral',   'emoji': '😐', 'hex': 'FF94A3B8'},
}

def _border():
    s = Side(style='thin', color='FFE2E8F0')
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val, bold=False, color='FF1E293B', bg=None,
          align='left', size=10, wrap=False, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = Font(name='Arial', bold=bold, color=color, size=size)
    c.border = _border()
    c.alignment = Alignment(
        horizontal=align, vertical='center', wrap_text=wrap)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def buat_excel(rows, filter_label):
    wb = Workbook()

    HDR_BG   = 'FF1E293B'
    HDR_FG   = 'FFFFFFFF'
    ALT_BG   = 'FFF8FAFC'
    BLUE_BG  = 'FF1E40AF'

    total = len(rows) or 1

    # ══ SHEET 1: RINGKASAN ════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Ringkasan'
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells('A1:G1')
    c = ws1['A1']
    c.value     = '📊 Laporan Monitoring Emosi Wajah Mahasiswa'
    c.font      = Font(name='Arial', bold=True, size=15, color='FF1E293B')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill      = PatternFill('solid', fgColor='FFE0F2FE')
    ws1.row_dimensions[1].height = 38

    ws1.merge_cells('A2:G2')
    ws1['A2'].value = filter_label
    ws1['A2'].font  = Font(name='Arial', size=9, color='FF64748B', italic=True)
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[2].height = 20
    ws1.row_dimensions[3].height = 8

    # Header tabel ringkasan
    for ci, h in enumerate(['Emosi','Emoji','Jumlah','%','Keterangan Otomatis'], 1):
        _cell(ws1, 4, ci, h, bold=True, color=HDR_FG, bg=HDR_BG, align='center')
    ws1.row_dimensions[4].height = 26

    # Hitung per emosi
    hitung = {k: 0 for k in EMOSI_META}
    angkat = 0
    for r in rows:
        e = r.get('emosi', 'neutral')
        if e in hitung: hitung[e] += 1
        if r.get('angkat_tangan'): angkat += 1

    sorted_emosi = sorted(hitung.items(), key=lambda x: x[1], reverse=True)

    KET = {
        'happy':   '✅ Mahasiswa senang & menikmati pembelajaran',
        'neutral': '📖 Fokus memperhatikan materi',
        'surprise':'🤔 Antusias / ada materi baru yang menarik',
        'sad':     '⚠️ Kurang bersemangat / ada kesulitan',
        'angry':   '⚠️ Frustrasi / tidak nyaman dengan materi',
        'fear':    '⚠️ Cemas / khawatir (mungkin ujian/tugas)',
        'disgust': '⚠️ Tidak tertarik pada materi',
    }

    for i, (k, jml) in enumerate(sorted_emosi):
        ri = 5 + i
        bg = ALT_BG if i % 2 == 0 else None
        pct = round(jml / total * 100, 2)
        _cell(ws1, ri, 1, EMOSI_META[k]['label'], bg=bg)
        _cell(ws1, ri, 2, EMOSI_META[k]['emoji'],  bg=bg, align='center')
        _cell(ws1, ri, 3, jml,                     bg=bg, align='center')
        _cell(ws1, ri, 4, pct,                     bg=bg, align='center', fmt='0.00"%"')
        _cell(ws1, ri, 5, KET.get(k,''),           bg=bg, wrap=True)
        ws1.row_dimensions[ri].height = 22

    row_tot = 5 + len(EMOSI_META)
    _cell(ws1, row_tot, 1, 'TOTAL', bold=True, color=HDR_FG, bg=BLUE_BG, align='center')
    c_tot = ws1.cell(row=row_tot, column=3,
                     value=f'=SUM(C5:C{row_tot-1})')
    c_tot.border  = _border()
    c_tot.font    = Font(name='Arial', bold=True)
    c_tot.alignment = Alignment(horizontal='center', vertical='center')
    _cell(ws1, row_tot, 4, '100%', bold=True, align='center')
    ws1.row_dimensions[row_tot].height = 24

    # Statistik tambahan
    rs = row_tot + 2
    ws1.merge_cells(f'A{rs}:G{rs}')
    _cell(ws1, rs, 1, '📈 Statistik Sesi', bold=True, color=HDR_FG, bg=HDR_BG)
    ws1.row_dimensions[rs].height = 26

    nama_mhs = list({r['nama'] for r in rows})
    dom_emosi = sorted_emosi[0][0] if sorted_emosi else 'neutral'
    stats = [
        ('Total Data Terdeteksi', len(rows)),
        ('Emosi Paling Dominan',  EMOSI_META[dom_emosi]['label']+' '+EMOSI_META[dom_emosi]['emoji']),
        ('Total Angkat Tangan',   angkat),
        ('Persentase Angkat Tangan', f'{round(angkat/total*100,1)}%'),
        ('Mahasiswa Terpantau',   ', '.join(nama_mhs) or '-'),
        ('Diekspor pada',         datetime.now().strftime('%d %B %Y %H:%M')),
    ]
    for j, (lbl, val) in enumerate(stats):
        r2 = rs + 1 + j
        _cell(ws1, r2, 1, lbl, bold=True)
        ws1.merge_cells(f'B{r2}:G{r2}')
        _cell(ws1, r2, 2, val)
        ws1.row_dimensions[r2].height = 20

    for ci, w in enumerate([18, 8, 14, 14, 40], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ══ SHEET 2: DATA LOG ═════════════════════════════════
    ws2 = wb.create_sheet('Data Log Lengkap')
    ws2.sheet_view.showGridLines = False

    cols2 = ['No','Waktu','Nama Mahasiswa','NIM','Kelas','Mata Kuliah',
             'Emosi','Emosi (ID)','Confidence (%)','Angkat Tangan']
    for ci, h in enumerate(cols2, 1):
        _cell(ws2, 1, ci, h, bold=True, color=HDR_FG, bg=HDR_BG, align='center')
    ws2.row_dimensions[1].height = 28

    for i, r in enumerate(rows):
        ri = i + 2
        bg = ALT_BG if i % 2 == 0 else None
        _cell(ws2, ri, 1,  i+1,                  bg=bg, align='center')
        _cell(ws2, ri, 2,  r.get('waktu',''),    bg=bg)
        _cell(ws2, ri, 3,  r.get('nama',''),     bg=bg)
        _cell(ws2, ri, 4,  r.get('nim',''),      bg=bg)   # dari join jika ada
        _cell(ws2, ri, 5,  r.get('kelas',''),    bg=bg)
        _cell(ws2, ri, 6,  r.get('mata_kuliah',''), bg=bg)
        _cell(ws2, ri, 7,  r.get('emosi',''),    bg=bg)
        em = r.get('emosi','neutral')
        c_em = _cell(ws2, ri, 8, r.get('label',''),
                     color='FFFFFFFF',
                     bg=EMOSI_META.get(em,{}).get('hex','FF94A3B8')[2:])
        _cell(ws2, ri, 9,  r.get('confidence',0), bg=bg, align='center', fmt='0.00')
        _cell(ws2, ri, 10,
              'Ya ✋' if r.get('angkat_tangan') else 'Tidak',
              bg=bg, align='center',
              color='FF22C55E' if r.get('angkat_tangan') else 'FF64748B')
        ws2.row_dimensions[ri].height = 18

    for ci, w in enumerate([5,20,22,14,12,20,10,14,14,14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ══ SHEET 3: ANALISIS PER MAHASISWA ═══════════════════
    ws3 = wb.create_sheet('Analisis per Mahasiswa')
    ws3.sheet_view.showGridLines = False

    emosi_keys = list(EMOSI_META.keys())
    cols3 = (['Nama','Kelas'] +
             [EMOSI_META[k]['label']+' '+EMOSI_META[k]['emoji'] for k in emosi_keys] +
             ['Angkat Tangan','Total','Emosi Dominan','Kesimpulan Otomatis'])
    for ci, h in enumerate(cols3, 1):
        _cell(ws3, 1, ci, h, bold=True, color=HDR_FG, bg=HDR_BG, align='center', wrap=True)
    ws3.row_dimensions[1].height = 34

    # Kelompokkan per mahasiswa
    per_mhs = {}
    for r in rows:
        key = r.get('nama','?')
        if key not in per_mhs:
            per_mhs[key] = {
                'kelas': r.get('kelas','-'),
                'counts': {k: 0 for k in emosi_keys},
                'angkat': 0, 'total': 0
            }
        e = r.get('emosi','neutral')
        if e in per_mhs[key]['counts']:
            per_mhs[key]['counts'][e] += 1
        if r.get('angkat_tangan'):
            per_mhs[key]['angkat'] += 1
        per_mhs[key]['total'] += 1

    for i, (nm, dat) in enumerate(per_mhs.items()):
        ri = i + 2
        bg = ALT_BG if i % 2 == 0 else None
        tot = dat['total'] or 1
        _cell(ws3, ri, 1, nm,          bold=True, bg=bg)
        _cell(ws3, ri, 2, dat['kelas'], bg=bg)
        for ci, k in enumerate(emosi_keys, 3):
            _cell(ws3, ri, ci, dat['counts'][k], bg=bg, align='center')
        ci_angkat = 3 + len(emosi_keys)
        ci_total  = ci_angkat + 1
        ci_dom    = ci_total  + 1
        ci_kes    = ci_dom    + 1
        _cell(ws3, ri, ci_angkat, dat['angkat'], bg=bg, align='center')
        ws3.cell(ri, ci_total).value  = f'=SUM(C{ri}:{get_column_letter(ci_angkat-1)}{ri})'
        ws3.cell(ri, ci_total).border = _border()
        ws3.cell(ri, ci_total).alignment = Alignment(horizontal='center',vertical='center')

        dom_k = max(emosi_keys, key=lambda k: dat['counts'][k])
        _cell(ws3, ri, ci_dom,
              EMOSI_META[dom_k]['label']+' '+EMOSI_META[dom_k]['emoji'],
              bold=True, align='center', bg=bg)

        # Kesimpulan otomatis
        happy_p  = dat['counts']['happy']  / tot * 100
        netral_p = dat['counts']['neutral'] / tot * 100
        angkat_p = dat['angkat'] / tot * 100
        sad_p    = dat['counts']['sad']   / tot * 100
        angry_p  = dat['counts']['angry'] / tot * 100

        if angkat_p >= 10:
            kes = '✅ Aktif bertanya — sangat antusias memahami materi'
        elif happy_p >= 35:
            kes = '✅ Senang & menikmati pembelajaran'
        elif happy_p + netral_p >= 65:
            kes = '📖 Fokus dan cukup memperhatikan materi'
        elif sad_p >= 30:
            kes = '⚠️ Terlihat kurang bersemangat'
        elif angry_p >= 20:
            kes = '⚠️ Menunjukkan ketidaknyamanan'
        else:
            kes = '📊 Ekspresi campuran, perlu perhatian lebih'

        _cell(ws3, ri, ci_kes, kes, bg=bg, wrap=True)
        ws3.row_dimensions[ri].height = 22

    for ci, w in enumerate([20,12]+[12]*len(emosi_keys)+[12,10,18,36], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
