"""
=============================================================
  MONITORING EMOSI WAJAH REAL-TIME v3.0
  Flask · OpenCV · DeepFace · MediaPipe

  Cara jalankan:
    pip install -r requirements.txt
    python app.py  →  http://localhost:5000
=============================================================
"""

import cv2, sqlite3, time, io, os, threading, base64
from datetime import datetime, timedelta
from collections import deque, Counter
from flask import (Flask, render_template, Response, jsonify,
                   request, send_file)
import numpy as np
from deepface import DeepFace
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── MediaPipe (opsional) ──────────────────
try:
    import mediapipe as mp
    _mp_pose = mp.solutions.pose
    _pose    = _mp_pose.Pose(
        static_image_mode=False, model_complexity=0,
        min_detection_confidence=0.55, min_tracking_confidence=0.55)
    MP_OK = True
    print("✅ MediaPipe aktif — deteksi angkat tangan tersedia")
except Exception:
    MP_OK = False
    print("⚠️  MediaPipe tidak tersedia (pip install mediapipe)")

app = Flask(__name__)
WAJAH_DB = 'data/wajah_db'
os.makedirs(WAJAH_DB, exist_ok=True)

# ── EMOSI CONFIG ─────────────────────────
EMOSI = {
    'angry':    {'label':'Marah',    'emoji':'😠', 'warna':'#ef4444', 'hex':'FFEF4444'},
    'disgust':  {'label':'Jijik',    'emoji':'🤢', 'warna':'#8b5cf6', 'hex':'FF8B5CF6'},
    'fear':     {'label':'Takut',    'emoji':'😨', 'warna':'#6366f1', 'hex':'FF6366F1'},
    'happy':    {'label':'Senang',   'emoji':'😊', 'warna':'#22c55e', 'hex':'FF22C55E'},
    'sad':      {'label':'Sedih',    'emoji':'😢', 'warna':'#3b82f6', 'hex':'FF3B82F6'},
    'surprise': {'label':'Terkejut', 'emoji':'😲', 'warna':'#f59e0b', 'hex':'FFF59E0B'},
    'neutral':  {'label':'Netral',   'emoji':'😐', 'warna':'#94a3b8', 'hex':'FF94A3B8'},
}

# ── Threshold confidence per emosi ──────────────────────────────────
# Tuning berdasarkan feedback:
#   angry/disgust/fear/surprise  → sangat kurang sensitif → threshold rendah
#   sad                          → terlalu sensitif       → threshold tinggi
#   happy / neutral              → sudah bagus
CONF_MIN = {
    'angry':    28,   # ↓ rendah — mudah terdeteksi
    'disgust':  25,   # ↓ rendah — paling susah, threshold sangat rendah
    'fear':     30,   # ↓ rendah
    'happy':    38,   # ✓ bagus
    'sad':      78,   # ↑ tinggi  — terlalu sensitif, perketat
    'surprise': 28,   # ↓ rendah
    'neutral':  18,   # ✓ fallback mudah
}

# Bobot koreksi: emosi yang under-detected diberi bonus skor sebelum threshold
BOOST = {
    'angry':   1.35,
    'disgust': 1.40,
    'fear':    1.30,
    'surprise':1.30,
    'happy':   1.0,
    'sad':     0.75,  # dikurangi agar tidak mudah muncul
    'neutral': 1.0,
}

def terapkan_threshold(dominan, semua):
    # Terapkan boost dulu, lalu cek threshold
    boosted = {k: v * BOOST.get(k, 1.0) for k, v in semua.items()}
    dom_boosted = max(boosted, key=boosted.get)

    # Jika emosi dominan setelah boost memenuhi threshold, pakai itu
    if boosted.get(dom_boosted, 0) >= CONF_MIN.get(dom_boosted, 50):
        return dom_boosted

    # Fallback: cari emosi lain yang memenuhi threshold
    for emo, pct in sorted(boosted.items(), key=lambda x: x[1], reverse=True):
        if pct >= CONF_MIN.get(emo, 50):
            return emo
    return 'neutral'

# ── DATABASE ─────────────────────────────
DB = 'data/log_emosi.db'

def get_db():
    con = sqlite3.connect(DB); con.row_factory = sqlite3.Row; return con

def init_db():
    os.makedirs('data', exist_ok=True)
    con = get_db()
    con.executescript('''
        CREATE TABLE IF NOT EXISTS siswa (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT NOT NULL, kelas TEXT DEFAULT '',
            nim TEXT DEFAULT '', foto TEXT, dibuat TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS log_emosi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            siswa_id INTEGER, mahasiswa TEXT NOT NULL DEFAULT 'Mahasiswa',
            kelas TEXT NOT NULL DEFAULT '-', mata_kuliah TEXT NOT NULL DEFAULT '-',
            emosi TEXT NOT NULL, label TEXT NOT NULL,
            confidence REAL NOT NULL DEFAULT 0,
            angkat_tgn INTEGER NOT NULL DEFAULT 0, waktu TEXT NOT NULL);
        CREATE INDEX IF NOT EXISTS idx_waktu ON log_emosi(waktu);
        CREATE INDEX IF NOT EXISTS idx_mapel ON log_emosi(mata_kuliah);
    ''')
    con.commit(); con.close()

def db_siswa_list():
    con=get_db(); rows=con.execute('SELECT * FROM siswa ORDER BY nama').fetchall(); con.close()
    return [dict(r) for r in rows]

def db_siswa_get(sid):
    con=get_db(); r=con.execute('SELECT * FROM siswa WHERE id=?',(sid,)).fetchone(); con.close()
    return dict(r) if r else None

def db_siswa_tambah(nama,kelas,nim):
    con=get_db()
    cur=con.execute('INSERT INTO siswa (nama,kelas,nim,dibuat) VALUES (?,?,?,?)',
        (nama,kelas,nim,datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    sid=cur.lastrowid; con.commit(); con.close(); return sid

def db_siswa_hapus(sid):
    s=db_siswa_get(sid)
    if s and s.get('foto') and os.path.exists(s['foto']):
        os.remove(s['foto'])
        for pkl in os.listdir(WAJAH_DB):
            if pkl.endswith('.pkl'): os.remove(os.path.join(WAJAH_DB,pkl))
    con=get_db(); con.execute('DELETE FROM siswa WHERE id=?',(sid,)); con.commit(); con.close()

def simpan_log(mahasiswa,kelas,mapel,emosi,label,conf,angkat=0,siswa_id=None):
    con=get_db()
    con.execute('''INSERT INTO log_emosi
        (siswa_id,mahasiswa,kelas,mata_kuliah,emosi,label,confidence,angkat_tgn,waktu)
        VALUES (?,?,?,?,?,?,?,?,?)''',
        (siswa_id,mahasiswa,kelas,mapel,emosi,label,
         round(conf,2),int(angkat),datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    con.commit(); con.close()

def query_log(limit=60,mapel=None,tgl_mulai=None,tgl_selesai=None):
    con=get_db(); sql='SELECT * FROM log_emosi WHERE 1=1'; args=[]
    if mapel and mapel!='semua': sql+=' AND mata_kuliah=?'; args.append(mapel)
    if tgl_mulai: sql+=' AND waktu>=?'; args.append(tgl_mulai)
    if tgl_selesai: sql+=' AND waktu<=?'; args.append(tgl_selesai+' 23:59:59')
    sql+=' ORDER BY id DESC'
    if limit: sql+=f' LIMIT {int(limit)}'
    rows=con.execute(sql,args).fetchall(); con.close()
    return [dict(r) for r in rows]

def daftar_mapel():
    con=get_db()
    rows=con.execute("SELECT DISTINCT mata_kuliah FROM log_emosi WHERE mata_kuliah NOT IN ('-','') ORDER BY mata_kuliah").fetchall()
    con.close(); return [r[0] for r in rows]

def hapus_semua_log():
    con=get_db(); con.execute('DELETE FROM log_emosi'); con.commit(); con.close()

# ── FACE RECOGNIZER (thread, ArcFace + distance filter) ─────────────
class FaceRecognizer:
    # Model & threshold jarak — makin kecil = makin mirip
    MODEL    = 'ArcFace'
    DIST_MAX = 0.45    # tolak jika jarak > nilai ini (0.0 = identik, 1.0 = beda)
    CONF_MIN_FR = 0.55 # confidence minimum DeepFace.find (jika tersedia)

    def __init__(self):
        self.lock     = threading.Lock()
        self._running = False
        self.nama     = None
        self.sid      = None
        self.foto     = None
        self.dist     = None   # jarak pengenalan (makin kecil = makin akurat)

    def kenali_async(self, frame_bgr):
        if self._running: return
        self._running = True
        threading.Thread(target=self._run, args=(frame_bgr.copy(),), daemon=True).start()

    @staticmethod
    def _preprocess(frame_bgr):
        """Tingkatkan kualitas frame sebelum pengenalan:
           - CLAHE untuk equalise pencahayaan
           - Resize ke 224x224 (input ArcFace)
           - Sharpening ringan
        """
        # Deteksi wajah dulu, crop area wajah agar lebih fokus
        cascade = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
        clf = cv2.CascadeClassifier(cascade)
        gray = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2GRAY)
        faces = clf.detectMultiScale(gray, 1.1, 5, minSize=(80, 80))
        if len(faces) > 0:
            x, y, w, h = faces[0]
            pad = int(max(w, h) * 0.20)
            x1 = max(0, x - pad); y1 = max(0, y - pad)
            x2 = min(frame_bgr.shape[1], x + w + pad)
            y2 = min(frame_bgr.shape[0], y + h + pad)
            frame_bgr = frame_bgr[y1:y2, x1:x2]

        # CLAHE pada channel L di LAB
        lab   = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))
        l     = clahe.apply(l)
        lab   = cv2.merge([l, a, b])
        enhanced = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)

        # Resize ke 224x224
        enhanced = cv2.resize(enhanced, (224, 224), interpolation=cv2.INTER_CUBIC)

        # Sharpening kernel
        kernel   = np.array([[0,-0.5,0],[-0.5,3,-0.5],[0,-0.5,0]])
        enhanced = cv2.filter2D(enhanced, -1, kernel)
        enhanced = np.clip(enhanced, 0, 255).astype(np.uint8)
        return enhanced

    def _run(self, frame):
        try:
            fotos = [f for f in os.listdir(WAJAH_DB)
                     if f.lower().endswith(('.jpg', '.jpeg', '.png'))
                     and not f.startswith('.')]
            if not fotos:
                with self.lock:
                    self.nama = None; self.sid = None
                    self.foto = None; self.dist = None
                return

            # Preprocess frame
            processed = self._preprocess(frame)
            tmp = 'data/_tmp_fr.jpg'
            cv2.imwrite(tmp, processed, [cv2.IMWRITE_JPEG_QUALITY, 95])

            hasil = DeepFace.find(
                img_path=tmp,
                db_path=WAJAH_DB,
                model_name=self.MODEL,
                distance_metric='cosine',
                enforce_detection=False,
                silent=True,
            )

            nama = None; sid = None; foto = None; dist = None

            if hasil and len(hasil) > 0:
                df = hasil[0]
                # Kolom jarak cosine dari ArcFace
                dist_col = [c for c in df.columns if 'distance' in c.lower() or 'cosine' in c.lower()]
                if not df.empty and dist_col:
                    df_sorted = df.sort_values(dist_col[0])
                    best_dist = float(df_sorted.iloc[0][dist_col[0]])

                    # Hanya terima jika jarak di bawah threshold
                    if best_dist <= self.DIST_MAX:
                        identity = df_sorted.iloc[0]['identity']
                        sid_str  = os.path.basename(identity).split('_')[0]
                        try:
                            sid  = int(sid_str)
                            s    = db_siswa_get(sid)
                            if s:
                                nama = s['nama']
                                foto = s.get('foto')
                                dist = round(best_dist, 3)
                        except: pass

            with self.lock:
                self.nama = nama; self.sid = sid
                self.foto = foto; self.dist = dist

        except Exception as e:
            pass
        finally:
            self._running = False
            try:
                if os.path.exists('data/_tmp_fr.jpg'):
                    os.remove('data/_tmp_fr.jpg')
            except: pass

    def get(self):
        with self.lock:
            return self.nama, self.sid, self.foto, self.dist

# ── DETEKTOR UTAMA ───────────────────────
class Detektor:
    SMOOTH   = 10    # window majority vote
    INTERVAL = 6     # analisis DeepFace tiap N frame
    WAJAH_FR = 30    # face recognition tiap N frame (lebih sering dari 60)
    SAVE_SEK = 5     # auto-save interval (detik)

    def __init__(self):
        self.kamera=None; self.aktif=False; self.lock=threading.Lock(); self.frame_no=0
        cascade=cv2.data.haarcascades+'haarcascade_frontalface_default.xml'
        self.cascade=cv2.CascadeClassifier(cascade)
        self.antrian=deque(maxlen=self.SMOOTH)
        self.emosi_kini='neutral'; self.label_kini='Netral'
        self.conf_kini=0.0; self.semua_kini={k:0.0 for k in EMOSI}
        self.angkat_tangan=False; self.angkat_ctr=0; self.ANGKAT_HOLD=45
        self.recognizer=FaceRecognizer()
        self.nama_dikenal=None; self.sid_dikenal=None
        self.mahasiswa='Mahasiswa'; self.kelas='-'; self.mata_kuliah='-'
        self.last_save=time.time()

    def buka(self,idx=0):
        if self.kamera and self.kamera.isOpened(): return True
        self.kamera=cv2.VideoCapture(idx)
        if self.kamera.isOpened():
            self.kamera.set(cv2.CAP_PROP_FRAME_WIDTH,640)
            self.kamera.set(cv2.CAP_PROP_FRAME_HEIGHT,480)
            self.aktif=True; return True
        return False

    def tutup(self):
        self.aktif=False
        if self.kamera: self.kamera.release(); self.kamera=None

    def _cek_angkat_tangan(self,frame_rgb):
        if not MP_OK: return False
        try:
            r=_pose.process(frame_rgb)
            if not r.pose_landmarks: return False
            lm=r.pose_landmarks.landmark; L=_mp_pose.PoseLandmark
            bahu_l=lm[L.LEFT_SHOULDER].y;  bahu_r=lm[L.RIGHT_SHOULDER].y
            tgn_l=lm[L.LEFT_WRIST].y;      tgn_r=lm[L.RIGHT_WRIST].y
            vis_l=lm[L.LEFT_WRIST].visibility>0.5
            vis_r=lm[L.RIGHT_WRIST].visibility>0.5
            return (vis_l and tgn_l<bahu_l-0.08) or (vis_r and tgn_r<bahu_r-0.08)
        except: return False

    def _analisis(self,frame):
        try:
            h=DeepFace.analyze(img_path=frame,actions=['emotion'],
                enforce_detection=False,detector_backend='opencv',silent=True)
            if isinstance(h,list): h=h[0]
            return h['dominant_emotion'], h['emotion']
        except: return 'neutral',{k:0.0 for k in EMOSI}

    def _smooth(self,dominan,semua):
        d=terapkan_threshold(dominan,semua)
        self.antrian.append(d)
        emo=Counter(self.antrian).most_common(1)[0][0]
        return emo, semua.get(emo,0)

    def _anotasi(self,frame,wajah,emosi,conf,angkat):
        info=EMOSI.get(emosi,EMOSI['neutral'])
        hx=info['warna'].lstrip('#')
        bgr=(int(hx[4:6],16),int(hx[2:4],16),int(hx[0:2],16))
        for (x,y,w,h) in wajah:
            cv2.rectangle(frame,(x,y),(x+w,y+h),bgr,2)
            teks=f"{info['label']} ({conf:.0f}%)"
            (tw,th),_=cv2.getTextSize(teks,cv2.FONT_HERSHEY_SIMPLEX,0.6,2)
            cv2.rectangle(frame,(x,y-th-14),(x+tw+8,y),bgr,-1)
            cv2.putText(frame,teks,(x+4,y-5),cv2.FONT_HERSHEY_SIMPLEX,0.6,(255,255,255),2)
        H=frame.shape[0]
        nama_disp=self.nama_dikenal or self.mahasiswa
        cv2.putText(frame,f"{nama_disp} | {self.mata_kuliah}",
            (10,H-14),cv2.FONT_HERSHEY_SIMPLEX,0.5,(200,200,200),1)
        if angkat:
            cv2.rectangle(frame,(8,6),(300,36),(20,160,60),-1)
            cv2.putText(frame,'Angkat Tangan / Bertanya',
                (14,27),cv2.FONT_HERSHEY_SIMPLEX,0.58,(255,255,255),2)
        return frame

    def generate(self):
        if not self.buka(): return
        dominan,semua='neutral',{k:0.0 for k in EMOSI}
        while self.aktif:
            ok,frame=self.kamera.read()
            if not ok: break
            self.frame_no+=1
            frame_rgb=cv2.cvtColor(frame,cv2.COLOR_BGR2RGB)
            gray=cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
            wajah=self.cascade.detectMultiScale(gray,1.1,5,minSize=(70,70))
            if MP_OK:
                angkat=self._cek_angkat_tangan(frame_rgb)
                if angkat: self.angkat_ctr=self.ANGKAT_HOLD
                else:
                    if self.angkat_ctr>0: self.angkat_ctr-=1
                self.angkat_tangan=(self.angkat_ctr>0)
            if self.frame_no%self.INTERVAL==0:
                dominan,semua=self._analisis(frame)
            emo_s,conf_s=self._smooth(dominan,semua)
            if self.frame_no%self.WAJAH_FR==0:
                self.recognizer.kenali_async(frame)
                n,sid,_,dist=self.recognizer.get()
                self.nama_dikenal=n; self.sid_dikenal=sid
            with self.lock:
                self.emosi_kini=emo_s; self.label_kini=EMOSI.get(emo_s,EMOSI['neutral'])['label']
                self.conf_kini=conf_s; self.semua_kini=semua
            now=time.time()
            if now-self.last_save>=self.SAVE_SEK:
                simpan_log(self.nama_dikenal or self.mahasiswa,self.kelas,self.mata_kuliah,
                    emo_s,EMOSI.get(emo_s,EMOSI['neutral'])['label'],
                    conf_s,angkat=int(self.angkat_tangan),siswa_id=self.sid_dikenal)
                self.last_save=now
            frame=self._anotasi(frame.copy(),wajah,emo_s,conf_s,self.angkat_tangan)
            _,buf=cv2.imencode('.jpg',frame,[cv2.IMWRITE_JPEG_QUALITY,85])
            yield b'--frame\r\nContent-Type: image/jpeg\r\n\r\n'+buf.tobytes()+b'\r\n'
        self.tutup()

    def get_status(self):
        with self.lock:
            n,_,foto,dist=self.recognizer.get()
            foto_url=None
            if foto and os.path.exists(foto):
                with open(foto,'rb') as f:
                    foto_url='data:image/jpeg;base64,'+base64.b64encode(f.read()).decode()
            # Hitung akurasi pengenalan dari jarak cosine (0=identik, 0.45=batas)
            akurasi_fr = None
            if dist is not None:
                akurasi_fr = round(max(0, (1 - dist / 0.45)) * 100, 1)
            return {
                'emosi':self.emosi_kini,'label':self.label_kini,
                'emoji':EMOSI.get(self.emosi_kini,EMOSI['neutral'])['emoji'],
                'conf':round(self.conf_kini,2),'angkat':self.angkat_tangan,
                'nama_dikenal':n,'foto_url':foto_url,
                'akurasi_fr': akurasi_fr,
                'semua':{k:{'label':EMOSI[k]['label'],'emoji':EMOSI[k]['emoji'],
                    'nilai':round(self.semua_kini.get(k,0),2),'warna':EMOSI[k]['warna']}
                    for k in EMOSI}
            }

cam=Detektor()

# ── HELPER: Preprocess foto siswa saat upload ──────────────────────
def _preprocess_foto_siswa(img_bgr):
    """Crop wajah + CLAHE enhance + resize ke 224x224.
       Hasilnya lebih bersih untuk ArcFace embeddings."""
    cascade = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
    clf = cv2.CascadeClassifier(cascade)
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    faces = clf.detectMultiScale(gray, 1.05, 4, minSize=(60, 60))
    if len(faces) > 0:
        x, y, w, h = max(faces, key=lambda r: r[2]*r[3])  # ambil wajah terbesar
        pad = int(max(w, h) * 0.25)
        x1=max(0,x-pad); y1=max(0,y-pad)
        x2=min(img_bgr.shape[1],x+w+pad); y2=min(img_bgr.shape[0],y+h+pad)
        img_bgr = img_bgr[y1:y2, x1:x2]
    # CLAHE
    lab = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2LAB)
    l,a,b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8,8))
    l = clahe.apply(l)
    img_bgr = cv2.cvtColor(cv2.merge([l,a,b]), cv2.COLOR_LAB2BGR)
    img_bgr = cv2.resize(img_bgr, (224, 224), interpolation=cv2.INTER_CUBIC)
    return img_bgr

# ── EXCEL EXPORT ─────────────────────────
def buat_excel(rows,judul,filter_info):
    wb=Workbook()
    thin=Side(style='thin',color='FFE2E8F0')
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal='center',vertical='center',wrap_text=True)
    lft=Alignment(horizontal='left',vertical='center')

    def hcell(ws,r,c,v,bg='1E293B',fc='FFFFFF',bold=True,al=None):
        cell=ws.cell(r,c,v); cell.fill=PatternFill('solid',fgColor=bg)
        cell.font=Font(name='Arial',bold=bold,color=fc,size=10)
        cell.border=bdr; cell.alignment=al or ctr; return cell

    def dcell(ws,r,c,v,bold=False,al=None,bg=None):
        cell=ws.cell(r,c,v); cell.font=Font(name='Arial',bold=bold,size=10)
        cell.border=bdr; cell.alignment=al or lft
        if bg: cell.fill=PatternFill('solid',fgColor=bg)
        return cell

    ALT='F8FAFC'

    # SHEET 1: RINGKASAN
    ws1=wb.active; ws1.title='Ringkasan'; ws1.sheet_view.showGridLines=False
    ws1.merge_cells('A1:I1'); ws1['A1'].value=f'📊 {judul}'
    ws1['A1'].font=Font(name='Arial',bold=True,size=15,color='1E293B')
    ws1['A1'].alignment=ctr; ws1['A1'].fill=PatternFill('solid',fgColor='F0F9FF')
    ws1.row_dimensions[1].height=38
    ws1.merge_cells('A2:I2'); ws1['A2'].value=filter_info
    ws1['A2'].font=Font(name='Arial',size=10,color='64748B',italic=True)
    ws1['A2'].alignment=ctr; ws1.row_dimensions[3].height=8

    total=len(rows) or 1; hitung={k:0 for k in EMOSI}; tgn=0; nm_set=set(); mp_set=set()
    for r in rows:
        if r['emosi'] in hitung: hitung[r['emosi']]+=1
        if r.get('angkat_tgn'): tgn+=1
        nm_set.add(r['mahasiswa']); mp_set.add(r['mata_kuliah'])

    for ci,h in enumerate(['Emosi','Emoji','Jumlah','Persen (%)','Keterangan Otomatis'],1):
        hcell(ws1,4,ci,h); ws1.row_dimensions[4].height=26

    ket_map={'happy':'✅ Mahasiswa senang & antusias belajar','neutral':'📖 Fokus dan memperhatikan materi',
        'surprise':'🤔 Antusias / materi baru menarik','sad':'⚠️ Terlihat kurang bersemangat',
        'fear':'⚠️ Cemas atau khawatir','angry':'⚠️ Tidak nyaman / frustrasi',
        'disgust':'⚠️ Tidak tertarik dengan materi'}

    for i,(k,jml) in enumerate(sorted(hitung.items(),key=lambda x:x[1],reverse=True)):
        ri=5+i; bg=ALT if i%2==0 else 'FFFFFF'
        dcell(ws1,ri,1,EMOSI[k]['label'],bg=bg); dcell(ws1,ri,2,EMOSI[k]['emoji'],al=ctr,bg=bg)
        dcell(ws1,ri,3,jml,al=ctr,bg=bg)
        c=dcell(ws1,ri,4,round(jml/total*100,2),al=ctr,bg=bg); c.number_format='0.00"%"'
        dcell(ws1,ri,5,ket_map.get(k,''),bg=bg); ws1.row_dimensions[ri].height=20

    rt=5+len(EMOSI); hcell(ws1,rt,1,'TOTAL',bg='1E40AF')
    ws1.cell(rt,3,f'=SUM(C5:C{rt-1})').border=bdr; ws1.cell(rt,4,'100%').border=bdr
    ws1.row_dimensions[rt].height=22
    dom_k=max(hitung,key=hitung.get)
    stats=[('Total Data Deteksi',len(rows)),
        ('Emosi Dominan',EMOSI[dom_k]['label']+' '+EMOSI[dom_k]['emoji']),
        ('Jumlah Angkat Tangan',tgn),('% Angkat Tangan',f'{round(tgn/total*100,1)}%'),
        ('Mahasiswa Terpantau',', '.join(sorted(nm_set)) or '-'),
        ('Mata Kuliah',', '.join(sorted(mp_set)) or '-'),
        ('Waktu Ekspor',datetime.now().strftime('%d %B %Y %H:%M'))]
    rs=rt+2; ws1.merge_cells(f'A{rs}:I{rs}'); hcell(ws1,rs,1,'📈 Statistik')
    ws1.row_dimensions[rs].height=24
    for j,(lbl,val) in enumerate(stats):
        r2=rs+1+j; dcell(ws1,r2,1,lbl,bold=True)
        ws1.merge_cells(f'B{r2}:I{r2}'); dcell(ws1,r2,2,val); ws1.row_dimensions[r2].height=18
    for ci,w in enumerate([18,7,14,14,38],1): ws1.column_dimensions[get_column_letter(ci)].width=w

    # SHEET 2: DATA LOG
    ws2=wb.create_sheet('Data Log'); ws2.sheet_view.showGridLines=False
    hdrs=['No','Waktu','Mahasiswa','Kelas','Mata Kuliah','Emosi','Label','Confidence (%)','Angkat Tgn']
    for ci,h in enumerate(hdrs,1): hcell(ws2,1,ci,h); ws2.row_dimensions[1].height=26
    for i,r in enumerate(rows):
        ri=i+2; bg=ALT if i%2==0 else 'FFFFFF'
        dcell(ws2,ri,1,i+1,al=ctr,bg=bg); dcell(ws2,ri,2,r['waktu'],bg=bg)
        dcell(ws2,ri,3,r['mahasiswa'],bg=bg); dcell(ws2,ri,4,r['kelas'],bg=bg)
        dcell(ws2,ri,5,r['mata_kuliah'],bg=bg); dcell(ws2,ri,6,r['emosi'],bg=bg)
        c=ws2.cell(ri,7,r['label']); c.border=bdr; c.alignment=ctr
        if r['emosi'] in EMOSI:
            c.fill=PatternFill('solid',fgColor=EMOSI[r['emosi']]['hex'][2:])
            c.font=Font(name='Arial',size=10,color='FFFFFF')
        cf=dcell(ws2,ri,8,r['confidence'],al=ctr,bg=bg); cf.number_format='0.00'
        dcell(ws2,ri,9,'Ya ✋' if r.get('angkat_tgn') else 'Tidak',al=ctr,bg=bg)
        ws2.row_dimensions[ri].height=16
    for ci,w in enumerate([5,20,18,12,20,10,12,14,14],1): ws2.column_dimensions[get_column_letter(ci)].width=w

    # SHEET 3: PER MAHASISWA
    ws3=wb.create_sheet('Per Mahasiswa'); ws3.sheet_view.showGridLines=False
    cols3=['Mahasiswa']+[EMOSI[k]['label'] for k in EMOSI]+['Angkat Tgn','Total','Dominan','Kesimpulan']
    for ci,h in enumerate(cols3,1): hcell(ws3,1,ci,h); ws3.row_dimensions[1].height=26
    per_mhs={}
    for r in rows:
        nm=r['mahasiswa']
        if nm not in per_mhs: per_mhs[nm]={k:0 for k in EMOSI}; per_mhs[nm]['_t']=0
        if r['emosi'] in per_mhs[nm]: per_mhs[nm][r['emosi']]+=1
        if r.get('angkat_tgn'): per_mhs[nm]['_t']+=1
    for i,(nm,dat) in enumerate(per_mhs.items()):
        ri=i+2; bg=ALT if i%2==0 else 'FFFFFF'
        tot=sum(dat[k] for k in EMOSI) or 1
        dcell(ws3,ri,1,nm,bold=True,bg=bg)
        for ci,k in enumerate(EMOSI,2): dcell(ws3,ri,ci,dat[k],al=ctr,bg=bg)
        ct=2+len(EMOSI); ctot=ct+1; cdom=ctot+1; ckes=cdom+1
        dcell(ws3,ri,ct,dat['_t'],al=ctr,bg=bg)
        dcell(ws3,ri,ctot,f'=SUM(B{ri}:{get_column_letter(ct-1)}{ri})',al=ctr,bg=bg)
        dk=max(EMOSI.keys(),key=lambda k:dat[k])
        c=ws3.cell(ri,cdom,EMOSI[dk]['label']+' '+EMOSI[dk]['emoji']); c.border=bdr; c.alignment=ctr
        c.fill=PatternFill('solid',fgColor=EMOSI[dk]['hex'][2:]); c.font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        hp=dat['happy']/tot*100; tp=dat['_t']/tot*100
        if hp>=40 or tp>=15: kes='✅ Sangat antusias & aktif bertanya'
        elif (hp+dat['neutral']/tot*100)>=60: kes='📖 Cukup fokus dan memperhatikan'
        elif dat['sad']/tot*100>=30: kes='⚠️ Kurang bersemangat, perlu perhatian'
        elif dat['angry']/tot*100>=20: kes='⚠️ Menunjukkan ketidaknyamanan'
        else: kes='📊 Ekspresi campuran / bervariasi'
        dcell(ws3,ri,ckes,kes,bg=bg); ws3.row_dimensions[ri].height=18
    for ci,w in enumerate([18]+[10]*len(EMOSI)+[10,10,18,35],1): ws3.column_dimensions[get_column_letter(ci)].width=w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── ROUTES ───────────────────────────────
@app.route('/')
def index():
    return render_template('index.html',emosi_list=EMOSI,mapel_list=daftar_mapel())

@app.route('/siswa')
def halaman_siswa():
    return render_template('mahasiswa.html',siswa_list=db_siswa_list())

@app.route('/video')
def video():
    return Response(cam.generate(),mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/status')
def api_status(): return jsonify(cam.get_status())

@app.route('/api/log')
def api_log(): return jsonify(query_log(limit=60,mapel=request.args.get('mapel')))

@app.route('/api/mapel')
def api_mapel(): return jsonify(daftar_mapel())

@app.route('/api/set-sesi',methods=['POST'])
def api_set_sesi():
    d=request.json
    cam.mahasiswa=(d.get('mahasiswa') or 'Mahasiswa').strip()
    cam.kelas=(d.get('kelas') or '-').strip()
    cam.mata_kuliah=(d.get('mata_kuliah') or '-').strip()
    return jsonify({'ok':True})

@app.route('/api/simpan-sekarang',methods=['POST'])
def api_simpan_sekarang():
    st=cam.get_status()
    simpan_log(cam.nama_dikenal or cam.mahasiswa,cam.kelas,cam.mata_kuliah,
        st['emosi'],st['label'],st['conf'],angkat=int(st['angkat']),siswa_id=cam.sid_dikenal)
    return jsonify({'ok':True})

@app.route('/api/hapus-log',methods=['POST'])
def api_hapus_log(): hapus_semua_log(); return jsonify({'ok':True})

@app.route('/api/stop')
def api_stop(): cam.tutup(); return jsonify({'ok':True})

@app.route('/api/ekspor-excel')
def api_ekspor_excel():
    f=request.args.get('filter','semua'); mapel=request.args.get('mapel','semua')
    now=datetime.now(); tgl_mulai=None; label='Semua Data'
    tgl_akhir=now.strftime('%Y-%m-%d')
    if f=='harian': tgl_mulai=now.strftime('%Y-%m-%d'); label=f'Hari Ini ({now.strftime("%d %B %Y")})'
    elif f=='mingguan': tgl_mulai=(now-timedelta(days=7)).strftime('%Y-%m-%d'); label='7 Hari Terakhir'
    elif f=='bulanan': tgl_mulai=(now-timedelta(days=30)).strftime('%Y-%m-%d'); label='30 Hari Terakhir'
    elif f=='tiga_bln': tgl_mulai=(now-timedelta(days=90)).strftime('%Y-%m-%d'); label='3 Bulan Terakhir'
    elif f=='enam_bln': tgl_mulai=(now-timedelta(days=180)).strftime('%Y-%m-%d'); label='6 Bulan (1 Semester)'
    if mapel and mapel!='semua': label+=f' · {mapel}'
    rows=query_log(limit=None,mapel=mapel if mapel!='semua' else None,tgl_mulai=tgl_mulai,tgl_selesai=tgl_akhir)
    if not rows: return jsonify({'error':'Tidak ada data untuk filter ini'}),404
    fi=f'Filter: {label}  |  Total: {len(rows)} data  |  Diekspor: {now.strftime("%d %B %Y %H:%M")}'
    buf=buat_excel(rows,'Laporan Monitoring Emosi Wajah',fi)
    nama=f'laporan_emosi_{f}_{now.strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,download_name=nama)

@app.route('/api/siswa')
def api_siswa(): return jsonify(db_siswa_list())

@app.route('/api/tambah-siswa',methods=['POST'])
def api_tambah_siswa():
    nama=(request.form.get('nama') or '').strip()
    kelas=(request.form.get('kelas') or '').strip()
    nim=(request.form.get('nim') or '').strip()
    if not nama: return jsonify({'ok':False,'pesan':'Nama wajib diisi'}),400
    sid=db_siswa_tambah(nama,kelas,nim)
    if 'foto' in request.files:
        f=request.files['foto']
        if f and f.filename:
            ext=os.path.splitext(f.filename)[1].lower() or '.jpg'
            fpath=os.path.join(WAJAH_DB,f"{sid}_{nama.replace(' ','_')}{ext}")
            # Simpan file asli dulu
            f.save(fpath)
            # Preprocess foto: crop wajah + enhance agar FR lebih akurat
            try:
                img = cv2.imread(fpath)
                if img is not None:
                    img = _preprocess_foto_siswa(img)
                    cv2.imwrite(fpath, img, [cv2.IMWRITE_JPEG_QUALITY, 95])
            except Exception as e:
                print(f"[WARN] Preprocess foto gagal: {e}")
            con=get_db(); con.execute('UPDATE siswa SET foto=? WHERE id=?',(fpath,sid)); con.commit(); con.close()
            # Hapus cache pkl DeepFace agar diregenerasi ulang dengan model ArcFace
            for pkl in os.listdir(WAJAH_DB):
                if pkl.endswith('.pkl'): os.remove(os.path.join(WAJAH_DB,pkl))
    return jsonify({'ok':True,'id':sid})

@app.route('/api/hapus-siswa/<int:sid>',methods=['POST'])
def api_hapus_siswa(sid): db_siswa_hapus(sid); return jsonify({'ok':True})

@app.route('/foto/<int:sid>')
def foto_siswa(sid):
    s=db_siswa_get(sid)
    if s and s.get('foto') and os.path.exists(s['foto']): return send_file(s['foto'],mimetype='image/jpeg')
    return '',404

if __name__=='__main__':
    init_db()
    print("="*50); print("  😊 MONITORING EMOSI WAJAH v3.0")
    print(f"  MediaPipe: {'✅ Aktif' if MP_OK else '❌ Tidak tersedia'}")
    print("  Buka: http://localhost:5000"); print("="*50)
    app.run(debug=False,host='0.0.0.0',port=5000,threaded=True)
